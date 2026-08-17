import os
import shutil
import sqlite3
import json
from datetime import datetime

import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserIconView
from kivy.core.window import Window
from kivy.utils import platform
from kivy.core.text import LabelBase, DEFAULT_FONT

# ---------- 尝试加载中文字体 ----------
font_candidates = [
    r"C:\Windows\Fonts\simhei.ttf",
    r"C:\Windows\Fonts\msyh.ttf",
    r"C:\Windows\Fonts\simkai.ttf",
    "simhei.ttf",  # 如果复制到项目目录
]
font_loaded = False
for font_path in font_candidates:
    if os.path.exists(font_path):
        try:
            LabelBase.register(DEFAULT_FONT, font_path)
            font_loaded = True
            break
        except:
            pass
if not font_loaded:
    print("警告：未找到中文字体，界面中文可能无法正常显示。请复制 simhei.ttf 到项目目录。")

# 尝试导入 plyer
try:
    from plyer import camera, filechooser
    HAS_PLYER = True
except ImportError:
    HAS_PLYER = False

# 尝试导入 openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------- 路径配置 ----------
if platform == "android":
    from android.storage import primary_external_storage_path
    APP_DATA_DIR = os.path.join(primary_external_storage_path(), "验房数据")
    os.makedirs(APP_DATA_DIR, exist_ok=True)
    DOWNLOAD_DIR = os.path.join(primary_external_storage_path(), "Download")
else:
    APP_DATA_DIR = os.getcwd()
    DOWNLOAD_DIR = os.getcwd()

SCREENSHOT_DIR = os.path.join(APP_DATA_DIR, "photos")
os.makedirs(SCREENSHOT_DIR, exist_ok=True)
DB_PATH = os.path.join(APP_DATA_DIR, "inspection.db")

# ---------- 数据库 ----------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS problems (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        building TEXT,
        room_no TEXT,
        description TEXT,
        category TEXT,
        nature TEXT,
        repair_unit TEXT,
        images TEXT,
        created_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    conn.commit()
    conn.close()

def get_setting(key, default=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else default

def save_setting(key, value):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def add_problem(building, room_no, description, category, nature, repair_unit, images):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO problems (building, room_no, description, category, nature, repair_unit, images, created_at) VALUES (?,?,?,?,?,?,?,?)",
              (building, room_no, description, category, nature, repair_unit, json.dumps(images), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_all_problems():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM problems ORDER BY created_at DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def delete_problem(pid):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT images FROM problems WHERE id=?", (pid,))
    row = c.fetchone()
    if row and row[0]:
        images = json.loads(row[0])
        for img in images:
            if os.path.exists(img):
                try:
                    os.remove(img)
                except:
                    pass
    c.execute("DELETE FROM problems WHERE id=?", (pid,))
    conn.commit()
    conn.close()

# ---------- 自动识别 ----------
PROBLEM_TYPES = {
    "空鼓": ["空鼓"],
    "裂缝": ["裂缝", "开裂", "裂纹"],
    "渗水": ["渗水", "漏水", "渗漏", "水渍", "水印"],
    "门窗": ["门窗", "窗户", "门", "窗框", "门框", "把手", "锁"],
    "电路": ["电路", "插座", "开关", "灯", "电线", "配电"],
    "水路": ["水路", "水管", "水龙头", "地漏", "下水", "排水"],
    "墙面": ["墙面", "墙皮", "墙砖", "墙纸", "抹灰"],
    "地面": ["地面", "地板", "地砖", "地坪"],
    "顶面": ["顶面", "天花板", "吊顶"],
}
LEVELS = {
    "严重": ["严重", "重大", "厉害", "很严重", "特别严重"],
    "一般": ["一般", "中等", "普通", "中度", "有点"],
    "轻微": ["轻微", "轻度", "小问题", "不严重", "些许"],
}

def parse_text(text):
    category = ""
    nature = ""
    for cat, kws in PROBLEM_TYPES.items():
        for kw in kws:
            if kw in text:
                category = cat
                break
        if category:
            break
    for nat, kws in LEVELS.items():
        for kw in kws:
            if kw in text:
                nature = nat
                break
        if nature:
            break
    return category, nature

# ---------- UI ----------
class MainScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation='vertical', padding=10, spacing=10, **kwargs)
        self.temp_images = []

        # 楼栋房号
        top = BoxLayout(orientation='horizontal', size_hint=(1, None), height=50, spacing=10)
        self.building_input = TextInput(text=get_setting("building"), hint_text="楼栋", multiline=False)
        self.room_input = TextInput(text=get_setting("room_no"), hint_text="房号", multiline=False)
        top.add_widget(self.building_input)
        top.add_widget(self.room_input)
        self.add_widget(top)

        # 问题描述
        self.desc_input = TextInput(hint_text="问题描述（一句话自动识别）", multiline=False)
        self.add_widget(self.desc_input)

        # 维修单位
        self.repair_input = TextInput(hint_text="维修单位（可选）", multiline=False)
        self.add_widget(self.repair_input)

        # 按钮行
        btn_box = BoxLayout(orientation='horizontal', size_hint=(1, None), height=50, spacing=10)
        btn_photo = Button(text="拍照/选图", on_press=self.pick_photo)
        btn_add = Button(text="添加问题", on_press=self.add_record)
        btn_export = Button(text="导出Excel", on_press=self.export_excel)
        btn_box.add_widget(btn_photo)
        btn_box.add_widget(btn_add)
        btn_box.add_widget(btn_export)
        self.add_widget(btn_box)

        # 照片预览标签
        self.photo_label = Label(text="未选择照片", size_hint=(1, None), height=30)
        self.add_widget(self.photo_label)

        # 问题列表标题
        self.add_widget(Label(text="问题列表", size_hint=(1, None), height=30, bold=True))

        # 列表滚动区域
        self.scroll = ScrollView()
        self.list_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        self.list_layout.bind(minimum_height=self.list_layout.setter('height'))
        self.scroll.add_widget(self.list_layout)
        self.add_widget(self.scroll)

        self.refresh_list()

    def pick_photo(self, instance):
        if platform == "android" and HAS_PLYER:
            content = BoxLayout(orientation='vertical', spacing=10)
            btn_camera = Button(text="拍照", size_hint=(1, None), height=50)
            btn_gallery = Button(text="从相册选择", size_hint=(1, None), height=50)
            content.add_widget(btn_camera)
            content.add_widget(btn_gallery)
            popup = Popup(title="选择图片来源", content=content, size_hint=(0.8, 0.4))
            btn_camera.bind(on_press=lambda x: self.take_photo(popup))
            btn_gallery.bind(on_press=lambda x: self.choose_gallery(popup))
            popup.open()
        else:
            self.choose_gallery(None)

    def take_photo(self, popup=None):
        if popup:
            popup.dismiss()
        if HAS_PLYER:
            try:
                file_path = camera.take_picture()
                if file_path:
                    self.add_image_path(file_path)
            except Exception as e:
                self.show_toast(f"拍照失败: {e}")

    def choose_gallery(self, popup=None):
        if popup:
            popup.dismiss()
        if HAS_PLYER:
            try:
                file_paths = filechooser.open_file(title="选择照片", filters=[["图片", "*.jpg", "*.jpeg", "*.png"]], multiple=True)
                if file_paths:
                    for fp in file_paths:
                        self.add_image_path(fp)
            except Exception as e:
                self.show_toast(f"选择失败: {e}")
        else:
            self.show_file_chooser()

    def show_file_chooser(self):
        content = BoxLayout(orientation='vertical')
        filechooser = FileChooserIconView(filters=["*.jpg", "*.jpeg", "*.png"])
        btn_select = Button(text="选择", size_hint=(1, None), height=50)
        content.add_widget(filechooser)
        content.add_widget(btn_select)
        popup = Popup(title="选择照片", content=content, size_hint=(0.9, 0.9))
        def select_files(instance):
            for path, selection in filechooser.selection:
                if os.path.isfile(path):
                    self.add_image_path(path)
            popup.dismiss()
        btn_select.bind(on_press=select_files)
        popup.open()

    def add_image_path(self, source_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        ext = os.path.splitext(source_path)[1]
        dest = os.path.join(SCREENSHOT_DIR, f"{timestamp}{ext}")
        shutil.copy2(source_path, dest)
        self.temp_images.append(dest)
        self.photo_label.text = f"已选择 {len(self.temp_images)} 张照片"

    def add_record(self, instance):
        building = self.building_input.text.strip()
        room = self.room_input.text.strip()
        desc = self.desc_input.text.strip()
        repair = self.repair_input.text.strip()

        if not building or not room:
            self.show_toast("请填写楼栋和房号")
            return
        if not desc:
            self.show_toast("请输入问题描述")
            return

        save_setting("building", building)
        save_setting("room_no", room)

        category, nature = parse_text(desc)
        add_problem(building, room, desc, category, nature, repair, self.temp_images.copy())

        self.desc_input.text = ""
        self.repair_input.text = ""
        self.temp_images.clear()
        self.photo_label.text = "未选择照片"
        self.refresh_list()
        self.show_toast("✅ 已添加")

    def refresh_list(self):
        self.list_layout.clear_widgets()
        problems = get_all_problems()
        if not problems:
            self.list_layout.add_widget(Label(text="暂无记录", size_hint_y=None, height=40))
        for p in problems:
            pid, building, room, desc, category, nature, repair, images_json, created = p
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=120, padding=5, spacing=2)
            card.add_widget(Label(text=f"{building} {room}  [{category or '未分类'}] [{nature or '未定'}]", size_hint_y=None, height=25, bold=True))
            card.add_widget(Label(text=f"问题: {desc}", size_hint_y=None, height=25))
            if repair:
                card.add_widget(Label(text=f"维修单位: {repair}", size_hint_y=None, height=20))
            images = json.loads(images_json) if images_json else []
            if images:
                img_path = images[0]
                if os.path.exists(img_path):
                    img = Image(source=img_path, size_hint=(None, None), size=(60, 60))
                    img.allow_stretch = True
                    card.add_widget(img)
            btn_del = Button(text="删除", size_hint=(None, None), size=(60, 30), pos_hint={'right': 1})
            btn_del.bind(on_press=lambda instance, pid=pid: self.delete_record(pid))
            card.add_widget(btn_del)
            self.list_layout.add_widget(card)

    def delete_record(self, pid):
        delete_problem(pid)
        self.refresh_list()
        self.show_toast("已删除")

    def export_excel(self, instance):
        problems = get_all_problems()
        if not problems:
            self.show_toast("暂无数据")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "验房问题"

        headers = ["序号", "楼栋", "房号", "问题描述", "类别", "性质", "维修单位", "问题照片"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        widths = [6, 10, 12, 40, 10, 8, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        max_img_col = 8
        for p in problems:
            images_json = p[7]
            if images_json:
                img_count = len(json.loads(images_json))
                max_img_col = max(max_img_col, 8 + img_count - 1)
        for col in range(8, max_img_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        img_success = 0
        for idx, p in enumerate(problems, 1):
            pid, building, room, desc, category, nature, repair, images_json, created = p
            row_idx = idx + 1
            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=building)
            ws.cell(row=row_idx, column=3, value=room)
            ws.cell(row=row_idx, column=4, value=desc)
            ws.cell(row=row_idx, column=5, value=category or "")
            ws.cell(row=row_idx, column=6, value=nature or "")
            ws.cell(row=row_idx, column=7, value=repair or "")

            images = json.loads(images_json) if images_json else []
            if images:
                ws.row_dimensions[row_idx].height = 80
                current_col = 8
                for img_path in images:
                    img_abs_path = os.path.abspath(img_path)
                    if os.path.exists(img_abs_path):
                        try:
                            img = XLImage(img_abs_path)
                            ratio = 70 / img.height
                            img.width = int(img.width * ratio)
                            img.height = 70
                            cell_addr = f"{get_column_letter(current_col)}{row_idx}"
                            ws.add_image(img, cell_addr)
                            current_col += 1
                            img_success += 1
                        except Exception as ex:
                            print(f"图片插入失败: {img_abs_path} - {ex}")

        file_name = f"验房问题_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(DOWNLOAD_DIR, file_name)
        wb.save(file_path)
        self.show_toast(f"✅ 导出成功，{img_success}张照片。\n保存到：{file_path}")

    def show_toast(self, msg):
        popup = Popup(title="提示", content=Label(text=msg), size_hint=(0.8, 0.3))
        popup.open()

class InspectionApp(App):
    def build(self):
        init_db()
        return MainScreen()

if __name__ == "__main__":
    InspectionApp().run()